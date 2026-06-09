import json
import re
from pathlib import Path
from io import BytesIO
import logging
import base64

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import load_workbook

from reddevil.core import get_settings
from reddevil.mail import MailParams, MailAttachment
from reddevil.mail.mail import sendEmailMessage

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/fide", tags=["fide"])

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "template_fide_form.xlsx"
TRANSLATIONS_PATH = BASE_DIR / "translations.json"

with open(TRANSLATIONS_PATH, "r", encoding="utf-8") as f:
    TRANSLATIONS = json.load(f)

FIDE_FIELDS = [
    ("fide_laws_followed", "Will FIDE Laws be followed?"),
    ("national_championship_143a", "National Championship 1.43a"),
    ("on_fide_calendar", "Is information put on FIDE Calendar?"),
    ("tournament_report", "Tournament report"),
    ("tournament_type", "Tournament type"),
    ("event_name", "Event Name"),
    ("city", "City"),
    ("country", "Country"),
    ("expected_players", "Expected number of players"),
    ("tournament_system", "Tournament system"),
    ("rounds_reported", "Number of Rounds reported"),
    ("multiple_round_days", "Number of multiple round days"),
    ("female_only", "Female players only"),
    ("start_date", "Start Date (YYYY-MM-DD)"),
    ("end_date", "End Date (YYYY-MM-DD)"),
    ("title_norms", "Title Norms available"),
    ("gm_wgm_norms", "GM/WGM Norms available"),
    ("chief_arbiter_name", "Chief Arbiter Name"),
    ("chief_arbiter_fide_id", "Chief Arbiter FIDE-ID"),
    ("dep_chief_arbiter1_name", "Deputy Chief Arbiter 1 Name"),
    ("dep_chief_arbiter1_fide_id", "Deputy Chief Arbiter 1 FIDE-ID"),
    ("dep_chief_arbiter2_name", "Deputy Chief Arbiter 2 Name"),
    ("dep_chief_arbiter2_fide_id", "Deputy Chief Arbiter 2 FIDE-ID"),
    ("kind_of_arbiters", "Kind of Arbiters"),
    ("arbiter1_name", "Arbiter 1 Name"),
    ("arbiter1_fide_id", "Arbiter 1 FIDE-ID"),
    ("arbiter2_name", "Arbiter 2 Name"),
    ("arbiter2_fide_id", "Arbiter 2 FIDE-ID"),
    ("arbiter3_name", "Arbiter 3 Name"),
    ("arbiter3_fide_id", "Arbiter 3 FIDE-ID"),
    ("arbiter4_name", "Arbiter 4 Name"),
    ("arbiter4_fide_id", "Arbiter 4 FIDE-ID"),
    ("chief_organizer_name", "Chief Organizer Name"),
    ("chief_organizer_fide_id", "Chief Organizer FIDE-ID"),
    ("organizer1_name", "Organizer 1 Name"),
    ("organizer1_fide_id", "Organizer 1 FIDE-ID"),
    ("organizer2_name", "Organizer 2 Name"),
    ("organizer2_fide_id", "Organizer 2 FIDE-ID"),
    ("organizer3_name", "Organizer 3 Name"),
    ("organizer3_fide_id", "Organizer 3 FIDE-ID"),
    ("time_control_code", "Time Control"),
    ("time_control_desc", "Time Control Description"),
    ("timectl_other_desc", "Time Control Description if it is not listed"),
    ("timectl1_moves", "Moves to first time control"),
    ("timectl1_minutes", "Minutes to first time control"),
    ("timectl1_inc_type", "Increment/Delay (1st control)"),
    ("timectl1_inc_seconds", "Seconds of Increment/Delay (1st control)"),
    ("timectl2_moves", "Moves to second time control"),
    ("timectl2_minutes", "Minutes to second time control"),
    ("timectl2_inc_type", "Increment/Delay (2nd control)"),
    ("timectl2_inc_seconds", "Seconds of Increment/Delay (2nd control)"),
    ("timectl_final_minutes", "Minutes for final session"),
    ("timectl_final_inc_type", "Increment/Delay (final)"),
    ("timectl_final_inc_seconds", "Seconds of Increment/Delay (final)"),
    ("max_rating", "Max Rating"),
    ("age_limit", "Age Limit"),
    ("age_limit_value", "Age Limit Value"),
    ("all_digital_clocks", "All Digital Clocks"),
    ("internet_tx", "Internet Transmission"),
    ("internet_tx_boards", "Enter number of boards"),
    ("tiebreak_method", "Tiebreak Method"),
    ("tiebreak_other", "If Tiebreak Method = 'Other', specify 'Other'"),
    ("software", "Software"),
    ("software_other", "If Software = 'Other', specify 'Other'"),
    ("software_version", "Version Software"),
    ("pgn_provided", "Will PGN be provided?"),
    ("contact_email", "Contact E-mail"),
    ("homepage", "Internet Homepage"),
    ("prize_fund", "Prize Fund in euros"),
    ("remarks", "Remarks"),
]

MANDATORY_ALWAYS = [
    "invoice_email",
    "invoice_clubnr",
    "fide_laws_followed",
    "national_championship_143a",
    "on_fide_calendar",
    "tournament_report",
    "tournament_type",
    "event_name",
    "city",
    "country",
    "expected_players",
    "tournament_system",
    "rounds_reported",
    "multiple_round_days",
    "female_only",
    "start_date",
    "end_date",
    "title_norms",
    "gm_wgm_norms",
    "chief_arbiter_name",
    "chief_arbiter_fide_id",
    "kind_of_arbiters",
    "chief_organizer_name",
    "chief_organizer_fide_id",
    "time_control_code",
    "time_control_desc",
    "all_digital_clocks",
    "tiebreak_method",
    "software",
    "software_version",
    "contact_email",
    "homepage",
]

LOOKUP_DATA = {
    "yes_no": ["Yes", "No"],
    "age_limit_options": ["None", "Under", "Over"],
    "inc_delay_options": ["Increment", "Delay"],
    "tiebreak_options": [],
    "software_options": [],
    "tournament_report_options": [],
    "kind_of_arbiters_options": [],
    "tournament_system_options": [],
    "fide_people": {},
    "fide_names": [],
    "time_control_types": [],
    "time_control_desc": {},
}


def load_lookup_values():
    global LOOKUP_DATA
    if not TEMPLATE_PATH.exists():
        logger.error(f"Template path not found: {TEMPLATE_PATH}")
        return

    wb = load_workbook(TEMPLATE_PATH, data_only=True)

    if "Tiebreak_Method" in wb.sheetnames:
        ws_tb = wb["Tiebreak_Method"]
        LOOKUP_DATA["tiebreak_options"] = [c.value for c in ws_tb["A"] if c.value]

    if "Software" in wb.sheetnames:
        ws_sw = wb["Software"]
        opts = [c.value for c in ws_sw["A"] if c.value]
        if "Swar" not in opts:
            if "Other" in opts:
                idx = opts.index("Other")
                opts.insert(idx, "Swar")
            else:
                opts.append("Swar")
        LOOKUP_DATA["software_options"] = opts

    if "Tournament_Report" in wb.sheetnames:
        ws_tr = wb["Tournament_Report"]
        LOOKUP_DATA["tournament_report_options"] = [
            c.value for c in ws_tr["A"] if c.value
        ]

    if "Kind_of_Arbiters" in wb.sheetnames:
        ws_ka = wb["Kind_of_Arbiters"]
        LOOKUP_DATA["kind_of_arbiters_options"] = [
            c.value for c in ws_ka["A"] if c.value
        ]

    if "Tournament_System" in wb.sheetnames:
        ws_ts = wb["Tournament_System"]
        LOOKUP_DATA["tournament_system_options"] = [
            c.value for c in ws_ts["A"] if c.value
        ]

    if "FIDE_ID" in wb.sheetnames:
        ws_id = wb["FIDE_ID"]
        for row in ws_id.iter_rows(min_row=2, max_col=3, values_only=True):
            name, fide_id, license_flag = row
            if not name:
                continue
            LOOKUP_DATA["fide_people"][name] = {
                "id": fide_id if fide_id is not None else "",
                "license": (license_flag or "").strip(),
            }
            LOOKUP_DATA["fide_names"].append(name)

    if "Time_Control" in wb.sheetnames:
        ws_tc = wb["Time_Control"]
        header_row = [c.value for c in ws_tc[1]]
        col_letters = ["B", "C", "D"]
        types = []
        for idx, col in enumerate(col_letters, start=1):
            if idx < len(header_row):
                label = header_row[idx]
                if label and label not in types:
                    types.append(label)
        LOOKUP_DATA["time_control_types"] = types
        for t, col in zip(types, col_letters):
            values = [c.value for c in ws_tc[col][1:] if c.value]
            LOOKUP_DATA["time_control_desc"][t] = values


load_lookup_values()


@router.get("/form-data")
def get_form_data():
    return {"translations": TRANSLATIONS, "lookups": LOOKUP_DATA}


def fill_workbook(form_data):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["FIDE Registration Form"]

    ws["B7"] = form_data.get("invoice_email", "")
    ws["B8"] = form_data.get("invoice_clubnr", "")

    start_row = 10
    for index, (field_key, _label) in enumerate(FIDE_FIELDS):
        ws[f"B{start_row + index}"] = form_data.get(field_key, "")

    if (
        form_data.get("tournament_report") == "New long tournament"
        and "Rounds_Long_Tournament" in wb.sheetnames
    ):
        ws_rounds = wb["Rounds_Long_Tournament"]
        ws_rounds["B1"] = form_data.get("event_name", "")
        for r in range(2, 150):
            ws_rounds[f"B{r}"] = None
            ws_rounds[f"C{r}"] = None
            if r > 12:
                ws_rounds[f"A{r}"] = None
        try:
            n_rounds = int(form_data.get("rounds_reported", 0))
        except ValueError:
            n_rounds = 0
        for i in range(1, n_rounds + 1):
            row = i + 1
            ws_rounds[f"A{row}"] = f"Round {i} Date"
            ws_rounds[f"B{row}"] = form_data.get(f"round{i}_date", "")
            ws_rounds[f"C{row}"] = form_data.get(f"round{i}_report", "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_int(value, field_label, errors, lang, min_value=None, max_value=None):
    t_msg = TRANSLATIONS.get(lang, TRANSLATIONS["en"])["messages"]
    if value is None or value == "":
        return None
    try:
        iv = int(value)
    except ValueError:
        logger.error(f"{field_label} must be an integer")
        errors.append(f"{field_label} {t_msg['field_must_be_integer']}")
        return None
    if min_value is not None and iv < min_value:
        logger.error(f"{field_label} must be at least {min}")
        errors.append(
            f"{field_label} {t_msg['field_must_be_at_least'].replace('{min}', str(min_value))}"
        )
    if max_value is not None and iv > max_value:
        logger.error(f"{field_label} must be at most {max}")
        errors.append(
            f"{field_label} {t_msg['field_must_be_at_most'].replace('{max}', str(max_value))}"
        )
    return iv


def parse_float(value, field_label, errors, lang, min_value=None):
    t_msg = TRANSLATIONS.get(lang, TRANSLATIONS["en"])["messages"]
    if value is None or value == "":
        return None
    try:
        fv = float(value)
    except ValueError:
        logger.error(f"{field_label} must be a number")
        errors.append(f"{field_label} {t_msg['field_must_be_number']}")
        return None
    if min_value is not None and fv < min_value:
        logger.error(f"{field_label} must be at least {min}")
        errors.append(
            f"{field_label} {t_msg['field_must_be_at_least'].replace('{min}', str(min_value))}"
        )
        return None
    return fv


def validate_form(form, lang):
    errors = []
    trans = TRANSLATIONS.get(lang, TRANSLATIONS["en"])
    t_msg = trans["messages"]
    t_fields = trans["fields"]

    if form.get("tournament_report") == "New long tournament":
        try:
            n_rounds = int(form.get("rounds_reported") or 0)
        except ValueError:
            n_rounds = 0
        reports = []
        for i in range(1, n_rounds + 1):
            rep_val = form.get(f"round{i}_report")
            if rep_val:
                reports.append(rep_val)
        if reports:
            form["multiple_round_days"] = str(len(reports) - len(set(reports)))
        else:
            form["multiple_round_days"] = "0"

    for key in MANDATORY_ALWAYS:
        if not form.get(key):
            label = t_fields.get(key, key)
            logger.error(f"value required for {label}")
            errors.append(f"{label} {t_msg['value_required']}")

    event_name = form.get("event_name", "")
    if event_name and not re.fullmatch(r"[A-Za-z0-9 ]+", event_name):
        logger.error(
            f"Event name: {event_name} can only contain letters, number or spaces"
        )
        errors.append(
            f"{t_fields.get('event_name', 'Event Name')} {t_msg['only_letters_numbers_spaces']}"
        )

    if form.get("fide_laws_followed") != "Yes":
        logger.error("Fide laws must be followed")
        errors.append(t_msg["fide_laws_error"])

    if form.get("tiebreak_method") == "Other" and not form.get("tiebreak_other"):
        logger.error("Cannot have empty tiebreak other")
        errors.append(
            f"{t_fields.get('tiebreak_other', 'tiebreak_other')} {t_msg['value_required']}"
        )

    if form.get("software") == "Other" and not form.get("software_other"):
        logger.error("Cannot have empty software other")
        errors.append(
            f"{t_fields.get('software_other', 'software_other')} {t_msg['value_required']}"
        )

    rounds_reported = parse_int(
        form.get("rounds_reported"),
        t_fields.get("rounds_reported", "Number of Rounds reported"),
        errors,
        lang,
        min_value=0,
        max_value=100,
    )
    parse_int(
        form.get("expected_players"),
        t_fields.get("expected_players", "Expected number of players"),
        errors,
        lang,
        min_value=1,
        max_value=2500,
    )
    parse_int(
        form.get("multiple_round_days"),
        t_fields.get("multiple_round_days", "Number of multiple round days"),
        errors,
        lang,
        min_value=0,
        max_value=100,
    )

    age_limit = form.get("age_limit")
    age_limit_value = form.get("age_limit_value")
    if age_limit and age_limit != "None":
        if not age_limit_value:
            dep_msg = (
                t_msg["value_required_for"]
                .replace("{dependency}", t_fields.get("age_limit", "Age Limit"))
                .replace("{value}", age_limit)
            )
            logger.error("Age limit value is required")
            errors.append(
                f"{t_fields.get('age_limit_value', 'Age Limit Value')} {dep_msg}"
            )
        else:
            parse_int(
                age_limit_value,
                t_fields.get("age_limit_value", "Age Limit Value"),
                errors,
                lang,
                min_value=0,
            )

    parse_int(
        form.get("max_rating"),
        t_fields.get("max_rating", "Max Rating"),
        errors,
        lang,
        min_value=0,
    )

    if form.get("internet_tx") == "Yes":
        boards = form.get("internet_tx_boards")
        if not boards:
            dep_msg = (
                t_msg["value_required_for"]
                .replace(
                    "{dependency}", t_fields.get("internet_tx", "Internet Transmission")
                )
                .replace("{value}", "Yes")
            )
            logger.error(f"{boards} is required")
            errors.append(
                f"{t_fields.get('internet_tx_boards', 'Enter number of boards')} {dep_msg}"
            )
        else:
            parse_int(
                boards,
                t_fields.get("internet_tx_boards", "Enter number of boards"),
                errors,
                lang,
                min_value=1,
            )

    if form.get("time_control_desc") == "Other":
        if not form.get("timectl_other_desc"):
            dep_msg = (
                t_msg["value_required_for"]
                .replace(
                    "{dependency}",
                    t_fields.get("time_control_desc", "Time Control Description"),
                )
                .replace("{value}", "Other")
            )
            logger.error("Time control description is required")
            errors.append(
                f"{t_fields.get('timectl_other_desc', 'Time Control Description if not listed')} {dep_msg}"
            )
        tc_fields = [
            (
                "timectl1_moves",
                t_fields.get("timectl1_moves", "Moves to first time control"),
            ),
            (
                "timectl1_minutes",
                t_fields.get("timectl1_minutes", "Minutes to first time control"),
            ),
            (
                "timectl1_inc_seconds",
                t_fields.get(
                    "timectl1_inc_seconds", "Seconds of Increment/Delay (1st control)"
                ),
            ),
            (
                "timectl2_moves",
                t_fields.get("timectl2_moves", "Moves to second time control"),
            ),
            (
                "timectl2_minutes",
                t_fields.get("timectl2_minutes", "Minutes to second time control"),
            ),
            (
                "timectl2_inc_seconds",
                t_fields.get(
                    "timectl2_inc_seconds", "Seconds of Increment/Delay (2nd control)"
                ),
            ),
            (
                "timectl_final_minutes",
                t_fields.get("timectl_final_minutes", "Minutes for final session"),
            ),
            (
                "timectl_final_inc_seconds",
                t_fields.get(
                    "timectl_final_inc_seconds", "Seconds of Increment/Delay (final)"
                ),
            ),
        ]
        for key, label in tc_fields:
            if not form.get(key):
                dep_msg = (
                    t_msg["value_required_for"]
                    .replace(
                        "{dependency}",
                        t_fields.get("time_control_desc", "Time Control Description"),
                    )
                    .replace("{value}", "Other")
                )
                logger.error(f"Time Control Description: {key} field missing")
                errors.append(f"{label} {dep_msg}")
            else:
                parse_int(form.get(key), label, errors, lang, min_value=0)
        for key in ["timectl1_inc_type", "timectl2_inc_type", "timectl_final_inc_type"]:
            if not form.get(key):
                dep_msg = (
                    t_msg["value_required_for"]
                    .replace(
                        "{dependency}",
                        t_fields.get("time_control_desc", "Time Control Description"),
                    )
                    .replace("{value}", "Other")
                )
                errors.append(f"{t_fields.get(key, key)} {dep_msg}")

    if form.get("tournament_report") == "New long tournament":
        n = rounds_reported if rounds_reported is not None else 0
        prev_date = None
        prev_idx = None
        for i in range(1, n + 1):
            date_key = f"round{i}_date"
            report_key = f"round{i}_report"
            date_val = form.get(date_key)
            if not date_val:
                errors.append(t_msg["round_date_required"].replace("{num}", str(i)))
            else:
                if prev_date and date_val < prev_date:
                    msg = (
                        t_msg["round_dates_order_error"]
                        .replace("{num}", str(i))
                        .replace("{prev_num}", str(prev_idx))
                    )
                    errors.append(msg)
                prev_date = date_val
                prev_idx = i
            if not form.get(report_key):
                errors.append(t_msg["round_report_required"].replace("{num}", str(i)))
            else:
                parse_int(
                    form.get(report_key),
                    t_msg["round_report_required"].replace("{num}", str(i)),
                    errors,
                    lang,
                    min_value=1,
                )

    parse_float(
        form.get("prize_fund"),
        t_fields.get("prize_fund", "Prize Fund in euros"),
        errors,
        lang,
        min_value=0.01,
    )
    return errors


@router.post("/generate")
async def generate_fide_form(locale: str, formdata: dict):
    locale = locale or "en"
    form = formdata.get("formdata", {})
    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="Template not found")

    form["fide_laws_followed"] = "Yes"
    if form.get("software") == "Swar":
        form["software"] = "Other"
        form["software_other"] = "Swar (with JaVaFo)"

    homepage = form.get("homepage", "").strip()
    if homepage and not (
        homepage.startswith("http://") or homepage.startswith("https://")
    ):
        form["homepage"] = "https://" + homepage

    errors = validate_form(form, locale)
    if errors:
        return JSONResponse(
            status_code=400, content={"success": False, "errors": errors}
        )

    filled = fill_workbook(form)
    event_name = form.get("event_name", "").strip()
    safe_event = re.sub(r"\s+", "_", event_name) or "Unknown"
    filename = f"Tournament_Registration_{safe_event}.xlsx"

    # Send email with FIDE registration Excel attachment
    excel_content = filled.getvalue()
    encoded_excel = base64.b64encode(excel_content)

    excel_attachment = MailAttachment(
        filename=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content_base64=encoded_excel,
    )

    settings = get_settings()
    sender_email = settings.EMAIL.get("sender", "noreply@frbe-kbsb-ksb.be")
    club_number = form.get("invoice_clubnr", "N/A")
    mail_subject = f"{event_name} - test form (Jorian)"

    mail_body = f"""
    <p>Beste,</p>
    <p>Hierbij vindt u het FIDE-registratieformulier voor het toernooi: <strong>{event_name}</strong>.</p>
    <p><strong>Clubnummer:</strong> {club_number}</p>
    <br>
    <p>Groetjes!</p>
    """

    mail_params = MailParams(
        locale=locale,
        receiver="fide@frbe-kbsb-ksb.be",
        sender=sender_email,
        subject=mail_subject,
        template=mail_body,
        attachments=[excel_attachment],
    )

    try:
        sendEmailMessage(mail_params)
        logger.info(
            f"FIDE Registration email sent to fide@frbe-kbsb-ksb.be from {sender_email}"
        )
    except Exception:
        logger.exception("Failed to send FIDE registration email")

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        iter([filled.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
